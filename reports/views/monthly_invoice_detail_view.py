from datetime import datetime
from decimal import Decimal

from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Sum
from django.http import HttpResponse
from django.views.generic import TemplateView

from invoices.models import Invoice
from reports.utils import format_currency


class MonthlyInvoiceDetailView(LoginRequiredMixin, TemplateView):
    """Detailed invoice listing by month with Excel export."""

    template_name = "reports/monthly_invoice_detail.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Get month and year from query params (default to current month)
        now = datetime.now()
        month = int(self.request.GET.get("month", now.month))
        year = int(self.request.GET.get("year", now.year))

        # Generate month/year options for the filter
        current_year = now.year
        years = list(range(current_year - 2, current_year + 2))  # Last 2 years + next year
        months = [
            {"value": 1, "name": "January"},
            {"value": 2, "name": "February"},
            {"value": 3, "name": "March"},
            {"value": 4, "name": "April"},
            {"value": 5, "name": "May"},
            {"value": 6, "name": "June"},
            {"value": 7, "name": "July"},
            {"value": 8, "name": "August"},
            {"value": 9, "name": "September"},
            {"value": 10, "name": "October"},
            {"value": 11, "name": "November"},
            {"value": 12, "name": "December"},
        ]

        # Filter invoices by month and year
        invoices = (
            Invoice.objects.filter(
                invoice_date__year=year,
                invoice_date__month=month,
            )
            .select_related("customer")
            .order_by("invoice_date", "invoice_no")
        )

        # Calculate totals
        total_amount = Decimal("0")
        total_paid = Decimal("0")
        total_due = Decimal("0")

        invoice_data = []
        for invoice in invoices:
            total_amount += invoice.total_amount
            total_paid += invoice.total_paid_amount
            total_due += invoice.total_due_amount

            invoice_data.append(
                {
                    "id": invoice.id,
                    "invoice_number": invoice.invoice_no,
                    "invoice_date": invoice.invoice_date,
                    "due_date": invoice.due_date,
                    "customer_name": invoice.customer.full_name,
                    "customer_id": invoice.customer.id,
                    "status": invoice.get_status_display(),
                    "status_code": invoice.status,
                    "total_amount": float(invoice.total_amount),
                    "total_amount_formatted": format_currency(invoice.total_amount),
                    "total_paid": float(invoice.total_paid_amount),
                    "total_paid_formatted": format_currency(invoice.total_paid_amount),
                    "total_due": float(invoice.total_due_amount),
                    "total_due_formatted": format_currency(invoice.total_due_amount),
                }
            )

        # Get month name
        month_name = months[month - 1]["name"]

        context.update(
            {
                "invoices": invoice_data,
                "selected_month": month,
                "selected_year": year,
                "month_name": month_name,
                "months": months,
                "years": years,
                "total_invoices": len(invoice_data),
                "total_amount": total_amount,
                "total_amount_formatted": format_currency(total_amount),
                "total_paid": total_paid,
                "total_paid_formatted": format_currency(total_paid),
                "total_due": total_due,
                "total_due_formatted": format_currency(total_due),
            }
        )

        return context

    def get(self, request, *args, **kwargs):
        # Check if Excel export is requested
        if request.GET.get("export") == "excel":
            return self.export_excel()
        return super().get(request, *args, **kwargs)

    def export_excel(self):
        """Export invoice data to Excel."""
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        # Get the same data as the view
        context = self.get_context_data()
        invoices = context["invoices"]
        month_name = context["month_name"]
        year = context["selected_year"]

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet()
        ws.title = f"{month_name} {year}"

        # Add title
        ws.merge_cells("A1:I1")
        title_cell = ws["A1"]
        title_cell.value = f"Invoice Report - {month_name} {year}"
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 30

        # Add headers
        headers = [
            "Invoice Number",
            "Invoice Date",
            "Due Date",
            "Customer",
            "Status",
            "Total Amount",
            "Total Paid",
            "Total Due",
        ]

        header_fill = PatternFill(start_color="0d6efd", end_color="0d6efd", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add data
        for row_num, invoice in enumerate(invoices, 4):
            ws.cell(row=row_num, column=1, value=invoice["invoice_number"])
            ws.cell(row=row_num, column=2, value=invoice["invoice_date"].strftime("%Y-%m-%d"))
            ws.cell(
                row=row_num, column=3, value=invoice["due_date"].strftime("%Y-%m-%d") if invoice["due_date"] else ""
            )
            ws.cell(row=row_num, column=4, value=invoice["customer_name"])
            ws.cell(row=row_num, column=5, value=invoice["status"])
            ws.cell(row=row_num, column=6, value=invoice["total_amount"])
            ws.cell(row=row_num, column=7, value=invoice["total_paid"])
            ws.cell(row=row_num, column=8, value=invoice["total_due"])

            # Format currency columns
            for col in [6, 7, 8]:
                ws.cell(row=row_num, column=col).number_format = "#,##0"

        # Add totals row
        total_row = len(invoices) + 4
        ws.cell(row=total_row, column=5, value="TOTAL").font = Font(bold=True)
        ws.cell(row=total_row, column=6, value=context["total_amount"]).font = Font(bold=True)
        ws.cell(row=total_row, column=7, value=context["total_paid"]).font = Font(bold=True)
        ws.cell(row=total_row, column=8, value=context["total_due"]).font = Font(bold=True)

        # Format total currency columns
        for col in [6, 7, 8]:
            ws.cell(row=total_row, column=col).number_format = "#,##0"

        # Auto-adjust column widths
        for col_num in range(1, 9):
            column_letter = get_column_letter(col_num)
            max_length = 0
            for cell in ws[column_letter]:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Create HTTP response
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = f'attachment; filename="invoices_{month_name}_{year}.xlsx"'

        wb.save(response)
        return response
