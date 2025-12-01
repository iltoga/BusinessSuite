import csv

import openpyxl
from django.contrib import admin
from django.http import HttpResponse
from openpyxl.utils import get_column_letter

from .models import Customer


class CustomerAdmin(admin.ModelAdmin):
    def delete_model(self, request, obj):
        can_delete, msg = obj.can_be_deleted(user=request.user)
        if not can_delete:
            from django.contrib import messages

            self.message_user(request, msg, messages.ERROR)
            return
        if msg:
            from django.contrib import messages

            self.message_user(request, msg, messages.WARNING)

        # Use force=True for superuser cascade delete
        force = request.user.is_superuser and obj.invoices.exists()
        obj.delete(force=force)

    list_display = ("full_name", "email", "telephone", "passport_number", "passport_expiration_date", "active")
    search_fields = ("first_name", "last_name", "email", "telephone", "passport_number")
    list_filter = ("active", "nationality")

    def export_as_csv(self, request, queryset):
        """Admin action: export selected customers to CSV including passport and birth place fields."""
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = "attachment; filename=customers_export.csv"
        writer = csv.writer(response)
        headers = [
            "Full Name",
            "Email",
            "Telephone",
            "Passport Number",
            "Passport Expiration Date",
            "Birth Place",
            "Active",
        ]
        writer.writerow(headers)
        for c in queryset:
            writer.writerow(
                [
                    c.full_name,
                    c.email or "",
                    c.telephone or "",
                    c.passport_number or "",
                    str(c.passport_expiration_date) if c.passport_expiration_date else "",
                    c.birth_place or "",
                    c.active,
                ]
            )
        return response

    export_as_csv.short_description = "Export selected customers as CSV (includes passports and birth place)"

    def export_as_excel(self, request, queryset):
        """Admin action: export selected customers to Excel including passport and birth place fields."""
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = [
            "Full Name",
            "Email",
            "Telephone",
            "Passport Number",
            "Passport Expiration Date",
            "Birth Place",
            "Active",
        ]
        for col_num, header in enumerate(headers, 1):
            ws.cell(row=1, column=col_num, value=header)
        for row_num, c in enumerate(queryset, start=2):
            ws.cell(row=row_num, column=1, value=c.full_name)
            ws.cell(row=row_num, column=2, value=c.email or "")
            ws.cell(row=row_num, column=3, value=c.telephone or "")
            ws.cell(row=row_num, column=4, value=c.passport_number or "")
            ws.cell(row=row_num, column=5, value=str(c.passport_expiration_date) if c.passport_expiration_date else "")
            ws.cell(row=row_num, column=6, value=c.birth_place or "")
            ws.cell(row=row_num, column=7, value=c.active)
        for col_num in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_num)].width = 25
        response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        response["Content-Disposition"] = "attachment; filename=customers_export.xlsx"
        wb.save(response)
        return response

    export_as_excel.short_description = "Export selected customers as Excel (includes passports and birth place)"

    actions = ["export_as_csv", "export_as_excel"]


admin.site.register(Customer, CustomerAdmin)
