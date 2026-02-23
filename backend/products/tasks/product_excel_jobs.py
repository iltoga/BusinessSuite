import os
import re
import traceback
from decimal import Decimal, InvalidOperation
from io import BytesIO

from core.models import AsyncJob
from core.services.logger_service import Logger
from core.services.push_notifications import PushNotificationService
from core.tasks.idempotency import acquire_task_lock, build_task_lock_key, release_task_lock
from django.contrib.auth import get_user_model
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.db import transaction
from django.utils import timezone
from huey.contrib.djhuey import db_task
from openpyxl import Workbook, load_workbook
from api.serializers.product_serializer import ProductCreateUpdateSerializer
from products.models import Product

logger = Logger.get_logger(__name__)
User = get_user_model()

EXPORT_COLUMNS: list[tuple[str, str]] = [
    ("code", "Code"),
    ("name", "Name"),
    ("description", "Description"),
    ("base_price", "Base Price"),
    ("retail_price", "Retail Price"),
]


def _safe_str(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_header(value) -> str:
    raw = _safe_str(value).lower()
    if not raw:
        return ""
    raw = raw.replace("(", " ").replace(")", " ")
    raw = raw.replace("-", " ").replace("/", " ")
    parts = [part for part in raw.split() if part]
    normalized = "_".join(parts)
    if normalized.endswith("_days"):
        normalized = normalized[: -len("_days")]
    return normalized


def _to_decimal_or_none(value, *, field_label: str) -> Decimal | None:
    if value is None:
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))

    raw = _safe_str(value)
    if raw == "":
        return None

    cleaned = re.sub(r"[^0-9,.\-]", "", raw)
    if cleaned == "":
        raise ValueError(f"Invalid {field_label}: {raw}")

    # Handle locale-like separators:
    # - "1.500.000,25" => "1500000.25"
    # - "1,500,000.25" => "1500000.25"
    if "," in cleaned and "." in cleaned:
        if cleaned.rfind(",") > cleaned.rfind("."):
            cleaned = cleaned.replace(".", "").replace(",", ".")
        else:
            cleaned = cleaned.replace(",", "")
    elif cleaned.count(",") == 1 and "." not in cleaned:
        left, right = cleaned.split(",", 1)
        if len(right) <= 2:
            cleaned = f"{left}.{right}"
        else:
            cleaned = f"{left}{right}"
    elif cleaned.count(".") > 1 and "," not in cleaned:
        cleaned = cleaned.replace(".", "")
    else:
        cleaned = cleaned.replace(",", "")

    try:
        return Decimal(cleaned)
    except (InvalidOperation, ValueError):
        raise ValueError(f"Invalid {field_label}: {raw}")


def _send_import_done_push(user, result: dict) -> None:
    if not user:
        return
    title = "Product import completed"
    body = (
        f"Created {result['created']} products, updated {result['updated']}, "
        f"errors {result['errors']}, skipped {result['skipped']}."
    )
    try:
        PushNotificationService().send_to_user(
            user=user,
            title=title,
            body=body,
            data={
                "type": "product-import",
                "created": str(result["created"]),
                "updated": str(result["updated"]),
                "errors": str(result["errors"]),
            },
            link="/products",
        )
    except Exception:
        logger.exception("Failed to send product import push notification to user #%s", getattr(user, "id", None))


@db_task()
def run_product_export_job(job_id: str, user_id: int | None = None, search_query: str = "") -> None:
    lock_key = build_task_lock_key(namespace="products_export_job", item_id=str(job_id))
    lock_token = acquire_task_lock(lock_key)
    if not lock_token:
        logger.warning("Product export task skipped due to lock contention: job_id=%s", job_id)
        return

    try:
        try:
            job = AsyncJob.objects.get(id=job_id)
        except AsyncJob.DoesNotExist:
            logger.error("AsyncJob %s not found for product export", job_id)
            return

        try:
            job.update_progress(5, "Preparing product export...", AsyncJob.STATUS_PROCESSING)
            queryset = Product.objects.all().order_by("name")
            query = (search_query or "").strip()
            if query:
                queryset = Product.objects.search_products(query).order_by("name")

            total = queryset.count()
            wb = Workbook()
            ws = wb.active
            ws.title = "Products"
            ws.append([column_label for _, column_label in EXPORT_COLUMNS])

            if total == 0:
                job.update_progress(80, "No products matched the export filter.")
            else:
                for index, product in enumerate(queryset.iterator(), 1):
                    retail_price = product.retail_price if product.retail_price is not None else product.base_price
                    ws.append(
                        [
                            product.code,
                            product.name,
                            product.description or "",
                            product.base_price if product.base_price is not None else Decimal("0.00"),
                            retail_price if retail_price is not None else Decimal("0.00"),
                        ]
                    )
                    if index == total or index % 10 == 0:
                        progress = min(90, 10 + int((index / total) * 80))
                        job.update_progress(progress, f"Exporting products... ({index}/{total})")

            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)

            timestamp = timezone.now().strftime("%Y%m%d_%H%M%S")
            filename = f"products_export_{timestamp}.xlsx"
            output_path = os.path.join("tmpfiles", "product_exports", str(job.id), filename)
            saved_path = default_storage.save(output_path, ContentFile(buffer.getvalue()))

            job.complete(
                result={
                    "file_path": saved_path,
                    "filename": filename,
                    "total_records": total,
                    "search_query": query,
                },
                message=f"Product export completed ({total} record(s)).",
            )
        except Exception as exc:
            logger.error("Product export job %s failed: %s", job_id, str(exc), exc_info=True)
            job.fail(str(exc), traceback.format_exc())
    finally:
        release_task_lock(lock_key, lock_token)


@db_task()
def run_product_import_job(job_id: str, user_id: int | None = None, file_path: str = "") -> None:
    lock_key = build_task_lock_key(namespace="products_import_job", item_id=str(job_id))
    lock_token = acquire_task_lock(lock_key)
    if not lock_token:
        logger.warning("Product import task skipped due to lock contention: job_id=%s", job_id)
        return

    try:
        try:
            job = AsyncJob.objects.get(id=job_id)
        except AsyncJob.DoesNotExist:
            logger.error("AsyncJob %s not found for product import", job_id)
            return

        user = User.objects.filter(id=user_id).first() if user_id else None

        try:
            job.update_progress(5, "Reading import file...", AsyncJob.STATUS_PROCESSING)

            with default_storage.open(file_path, "rb") as file_handle:
                workbook = load_workbook(file_handle, data_only=True)
            worksheet = workbook.active

            header_cells = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
            headers = [_normalize_header(value) for value in header_cells]
            if not headers:
                raise ValueError("Import file is missing a header row.")

            required_columns = {"code", "name"}
            missing_required = [column for column in required_columns if column not in headers]
            if missing_required:
                raise ValueError(f"Missing required column(s): {', '.join(missing_required)}")

            valid_columns = {column for column, _label in EXPORT_COLUMNS}
            rows = list(worksheet.iter_rows(min_row=2, values_only=True))
            total_rows = len(rows)

            created_count = 0
            updated_count = 0
            error_count = 0
            skipped_count = 0
            row_errors: list[dict] = []

            if total_rows == 0:
                job.update_progress(90, "Import file has no data rows.")

            for row_index, row_values in enumerate(rows, start=2):
                mapped = {headers[idx]: row_values[idx] for idx in range(min(len(headers), len(row_values)))}

                if all(_safe_str(value) == "" for value in mapped.values()):
                    skipped_count += 1
                    continue

                code = _safe_str(mapped.get("code"))
                if not code:
                    error_count += 1
                    row_errors.append({"row": row_index, "error": "Product code is required."})
                    continue

                name = _safe_str(mapped.get("name"))
                if not name:
                    error_count += 1
                    row_errors.append({"row": row_index, "code": code, "error": "Product name is required."})
                    continue

                try:
                    # Build payload only with fields present in file and supported by import template.
                    product_payload = {}
                    if "name" in mapped:
                        product_payload["name"] = name
                    if "description" in mapped:
                        product_payload["description"] = _safe_str(mapped.get("description"))
                    if "base_price" in mapped:
                        base_price = _to_decimal_or_none(mapped.get("base_price"), field_label="base_price")
                        if base_price is not None:
                            product_payload["base_price"] = str(base_price)
                    if "retail_price" in mapped:
                        retail_price = _to_decimal_or_none(mapped.get("retail_price"), field_label="retail_price")
                        if retail_price is not None:
                            product_payload["retail_price"] = str(retail_price)

                    # Explicitly ignore tasks/system/unrecognized columns from custom files.
                    for key in list(mapped.keys()):
                        if key not in valid_columns:
                            if key == "tasks":
                                logger.info("Ignoring tasks column in product import row %s", row_index)
                            continue

                    with transaction.atomic():
                        product = Product.objects.filter(code=code).first()
                        if product:
                            # PATCH semantics: only update fields provided in import file.
                            serializer = ProductCreateUpdateSerializer(
                                product,
                                data=product_payload,
                                partial=True,
                                context={"request": None},
                            )
                            serializer.is_valid(raise_exception=True)
                            serializer.save(updated_by=user)
                            updated_count += 1
                        else:
                            create_payload = {
                                "code": code,
                                "name": name,
                                "description": _safe_str(mapped.get("description")),
                            }
                            if "base_price" in product_payload:
                                create_payload["base_price"] = product_payload["base_price"]
                            if "retail_price" in product_payload:
                                create_payload["retail_price"] = product_payload["retail_price"]
                            serializer = ProductCreateUpdateSerializer(
                                data=create_payload,
                                context={"request": None},
                            )
                            serializer.is_valid(raise_exception=True)
                            serializer.save(created_by=user, updated_by=user)
                            created_count += 1
                except Exception as exc:
                    error_count += 1
                    row_errors.append({"row": row_index, "code": code, "error": str(exc)})

                processed_rows = row_index - 1
                if total_rows > 0 and (processed_rows == total_rows or processed_rows % 5 == 0):
                    progress = min(95, 10 + int((processed_rows / total_rows) * 85))
                    job.update_progress(progress, f"Importing products... ({processed_rows}/{total_rows})")

            result = {
                "total_rows": total_rows,
                "created": created_count,
                "updated": updated_count,
                "errors": error_count,
                "skipped": skipped_count,
                "row_errors": row_errors[:200],
            }
            job.complete(
                result=result,
                message=(
                    f"Product import completed: {created_count} created, {updated_count} updated, "
                    f"{error_count} error(s), {skipped_count} skipped."
                ),
            )
            _send_import_done_push(user, result)
        except Exception as exc:
            logger.error("Product import job %s failed: %s", job_id, str(exc), exc_info=True)
            job.fail(str(exc), traceback.format_exc())
        finally:
            if file_path:
                try:
                    if default_storage.exists(file_path):
                        default_storage.delete(file_path)
                except Exception:
                    logger.warning("Could not remove temporary import file: %s", file_path, exc_info=True)
    finally:
        release_task_lock(lock_key, lock_token)
