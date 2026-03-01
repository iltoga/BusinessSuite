from decimal import Decimal
from io import BytesIO
from tempfile import TemporaryDirectory
from unittest.mock import patch

from core.models import AsyncJob
from django.core.files.base import ContentFile
from django.core.files.storage import FileSystemStorage
from django.test import TestCase
from openpyxl import Workbook, load_workbook
from products.models import Product
from products.tasks.product_excel_jobs import run_product_export_job, run_product_import_job


def _run_task(task, **kwargs):
    if hasattr(task, "call_local"):
        return task.call_local(**kwargs)
    if hasattr(task, "func"):
        return task.func(**kwargs)
    return task(**kwargs)


class ProductExcelJobsTests(TestCase):
    def setUp(self):
        super().setUp()
        self.temp_dir = TemporaryDirectory()
        self.storage = FileSystemStorage(location=self.temp_dir.name, base_url="/media/")

    def tearDown(self):
        self.temp_dir.cleanup()
        super().tearDown()

    def test_export_includes_base_and_retail_prices(self):
        Product.objects.create(
            code="EXP-1",
            name="Export Product",
            description="Export description",
            base_price=Decimal("1500000.00"),
            retail_price=Decimal("1800000.00"),
        )
        job = AsyncJob.objects.create(task_name="products_export_excel", status=AsyncJob.STATUS_PENDING)

        with (
            patch("products.tasks.product_excel_jobs.default_storage", self.storage),
            patch("products.tasks.product_excel_jobs.acquire_task_lock", return_value="token-export"),
            patch("products.tasks.product_excel_jobs.release_task_lock"),
        ):
            _run_task(run_product_export_job, job_id=str(job.id), search_query="")

        job.refresh_from_db()
        self.assertEqual(job.status, AsyncJob.STATUS_COMPLETED)
        self.assertIsNotNone(job.result)
        file_path = None
        if job.result is not None:
            file_path = job.result.get("file_path")
        self.assertIsNotNone(file_path)

        with self.storage.open(file_path, "rb") as fh:
            workbook = load_workbook(fh, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))

        self.assertEqual(rows[0], ("Code", "Name", "Description", "Base Price", "Retail Price"))
        self.assertEqual(rows[1][0], "EXP-1")
        self.assertEqual(rows[1][1], "Export Product")
        self.assertEqual(rows[1][2], "Export description")
        self.assertEqual(Decimal(str(rows[1][3])), Decimal("1500000"))
        self.assertEqual(Decimal(str(rows[1][4])), Decimal("1800000"))

    def test_import_reads_base_and_retail_prices(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Code", "Name", "Description", "Base Price", "Retail Price"])
        sheet.append(["IMP-1", "Imported Product", "Imported description", "2000000", "2600000"])
        # Retail omitted -> must default to base price.
        sheet.append(["IMP-2", "Imported Product 2", "Imported description 2", "3000000", ""])

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        import_path = self.storage.save("tmpfiles/product_imports/products.xlsx", ContentFile(buffer.getvalue()))
        job = AsyncJob.objects.create(task_name="products_import_excel", status=AsyncJob.STATUS_PENDING)

        with (
            patch("products.tasks.product_excel_jobs.default_storage", self.storage),
            patch("products.tasks.product_excel_jobs.acquire_task_lock", return_value="token-import"),
            patch("products.tasks.product_excel_jobs.release_task_lock"),
            patch("products.tasks.product_excel_jobs._send_import_done_push"),
        ):
            _run_task(run_product_import_job, job_id=str(job.id), file_path=import_path)

        job.refresh_from_db()
        self.assertEqual(job.status, AsyncJob.STATUS_COMPLETED)
        self.assertIsNotNone(job.result)
        # mypy can't see the assertIsNotNone above, so copy to a local variable
        result = job.result or {}
        self.assertEqual(result["created"], 2)
        self.assertEqual(result["errors"], 0)

        first = Product.objects.get(code="IMP-1")
        second = Product.objects.get(code="IMP-2")

        self.assertEqual(first.base_price, Decimal("2000000.00"))
        self.assertEqual(first.retail_price, Decimal("2600000.00"))
        self.assertEqual(second.base_price, Decimal("3000000.00"))
        self.assertEqual(second.retail_price, Decimal("3000000.00"))
